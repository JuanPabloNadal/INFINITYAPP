"""Exportación de reportes a Excel (openpyxl) y PDF (reportlab)."""
import os
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
)
from reportlab.lib.utils import ImageReader

from ..models import REP_INFINITY, COMISION_PORCENTAJE
from ..utils.formato import (
    formato_moneda, formato_fecha, formato_porcentaje, simbolo_moneda,
)
from .reportes import balance_por_moneda

# Paleta de marca Infinity (carbón + naranja sobre blanco)
AZUL = "1D1D1F"        # carbón (encabezados oscuros)
NARANJA = "E87722"
GRIS_CLARO = "F4F4F6"

_IMG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "img")
LOGO_PDF = os.path.join(_IMG_DIR, "infinity-logo.png")


def _f(valor):
    return float(valor or 0)


def _comision_desc(linea):
    if linea.comision_tipo == COMISION_PORCENTAJE:
        return formato_porcentaje(linea.comision_porcentaje)
    return "Monto fijo"


# --------------------------------------------------------------------------
# EXCEL
# --------------------------------------------------------------------------
def exportar_excel(operaciones, titulo="Reporte de operaciones", subtitulo=""):
    wb = Workbook()

    # ---- Hoja Detalle ----
    ws = wb.active
    ws.title = "Detalle"

    encabezados = [
        "Fecha", "Tipo", "Propiedad", "Moneda", "Punta", "Agente",
        "Base de cálculo", "Comisión", "Retención %",
        "Comisión bruta", "Monto inmobiliaria", "Monto agente",
    ]

    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    if subtitulo:
        ws.append([subtitulo])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(encabezados))
        ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.append([])

    fila_header = ws.max_row + 1
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=fila_header, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for op in operaciones:
        lineas_inf = [l for l in op.lineas if l.genera_comision]
        if not lineas_inf:
            # Operación sin comisión para Infinity: fila informativa.
            ws.append([
                formato_fecha(op.fecha), op.tipo.capitalize(), op.propiedad,
                op.moneda, "Sin punta Infinity", "—",
                None, None, None, 0, 0, 0,
            ])
            continue
        for l in lineas_inf:
            ws.append([
                formato_fecha(op.fecha),
                op.tipo.capitalize(),
                op.propiedad,
                op.moneda,
                l.etiqueta_rol,
                l.agente.nombre_completo if l.agente else "Sin agente",
                _f(l.base_calculo),
                _comision_desc(l),
                _f(l.retencion_porcentaje),
                _f(l.comision_bruta),
                _f(l.monto_inmobiliaria),
                _f(l.monto_agente),
            ])
            # Figuras adicionales de esta línea (datero/captador de desarrollo/AAA).
            for fig in l.figuras:
                ws.append([
                    formato_fecha(op.fecha),
                    op.tipo.capitalize(),
                    op.propiedad,
                    op.moneda,
                    f"↳ {fig.etiqueta_tipo}",
                    fig.agente.nombre_completo if fig.agente else "Sin agente",
                    None,
                    formato_porcentaje(fig.porcentaje),
                    _f(fig.retencion_porcentaje),
                    _f(fig.comision_bruta),
                    _f(fig.monto_inmobiliaria),
                    _f(fig.monto_agente),
                ])

    # Formato de números (columnas de dinero y base)
    money_cols = [7, 10, 11, 12]
    for row in ws.iter_rows(min_row=fila_header + 1, max_row=ws.max_row):
        for col in money_cols:
            row[col - 1].number_format = '#,##0'

    _autoajustar(ws, len(encabezados))

    # ---- Hoja Resumen ----
    _hoja_resumen(wb, operaciones, titulo, subtitulo)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


def _hoja_resumen(wb, operaciones, titulo, subtitulo):
    ws = wb.create_sheet("Resumen")
    balance = balance_por_moneda(operaciones)

    ws.append([titulo + " — Resumen"])
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    if subtitulo:
        ws.append([subtitulo])
        ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws.append([])

    for moneda, datos in balance.items():
        fila = ws.max_row + 1
        ws.append([f"Moneda: {moneda}"])
        ws.cell(row=fila, column=1).font = Font(bold=True, size=12, color=NARANJA)

        ws.append(["Operaciones", datos["cantidad_operaciones"]])
        ws.append(["Comisión bruta total", _f(datos["comision_bruta"])])
        ws.append(["Total inmobiliaria", _f(datos["monto_inmobiliaria"])])
        ws.append(["Total agentes", _f(datos["monto_agente_total"])])
        for r in range(ws.max_row - 2, ws.max_row + 1):
            ws.cell(row=r, column=2).number_format = '#,##0'

        ws.append([])
        fila_h = ws.max_row + 1
        ws.append(["Agente", "Monto que se llevó"])
        for col in (1, 2):
            cc = ws.cell(row=fila_h, column=col)
            cc.font = Font(bold=True, color="FFFFFF")
            cc.fill = PatternFill("solid", fgColor=AZUL)
        for nombre, monto in datos["por_agente"]:
            ws.append([nombre, _f(monto)])
            ws.cell(row=ws.max_row, column=2).number_format = '#,##0'
        ws.append([])
        ws.append([])

    _autoajustar(ws, 2, minimo=18)


def _autoajustar(ws, n_cols, minimo=10):
    for col in range(1, n_cols + 1):
        letra = get_column_letter(col)
        largo = minimo
        for cell in ws[letra]:
            try:
                if cell.value is not None:
                    largo = max(largo, len(str(cell.value)) + 2)
            except Exception:
                pass
        ws.column_dimensions[letra].width = min(largo, 45)


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------
def exportar_pdf(operaciones, titulo="Reporte de operaciones", subtitulo=""):
    salida = BytesIO()
    doc = SimpleDocTemplate(
        salida, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=titulo,
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloInfinity", parent=estilos["Title"], fontSize=18,
        textColor=colors.HexColor("#" + AZUL), spaceAfter=2, alignment=0,
    )
    estilo_marca = ParagraphStyle(
        "Marca", parent=estilos["Normal"], fontSize=10,
        textColor=colors.HexColor("#666666"), spaceAfter=8,
    )
    estilo_sub = ParagraphStyle(
        "Sub", parent=estilos["Normal"], fontSize=10,
        textColor=colors.HexColor("#333333"), spaceAfter=10,
    )
    estilo_seccion = ParagraphStyle(
        "Seccion", parent=estilos["Heading2"], fontSize=12,
        textColor=colors.HexColor("#" + NARANJA), spaceBefore=10, spaceAfter=4,
    )

    elementos = []
    # Logo oficial en el encabezado (con respaldo a texto si falta el archivo).
    if os.path.exists(LOGO_PDF):
        try:
            ir = ImageReader(LOGO_PDF)
            iw, ih = ir.getSize()
            alto = 20 * mm
            ancho = alto * iw / ih
            logo = Image(LOGO_PDF, width=ancho, height=alto)
            logo.hAlign = "LEFT"
            elementos.append(logo)
            elementos.append(Spacer(1, 4))
        except Exception:
            elementos.append(Paragraph("INFINITY INMOBILIARIA", estilo_titulo))
    else:
        elementos.append(Paragraph("INFINITY INMOBILIARIA", estilo_titulo))
    elementos.append(Paragraph(titulo, estilo_sub))
    if subtitulo:
        elementos.append(Paragraph(subtitulo, estilo_sub))

    # Tabla de operaciones
    encabezados = [
        "Fecha", "Tipo", "Propiedad", "Mon.", "Punta", "Agente",
        "Com. bruta", "Inmob.", "Agente $",
    ]
    data = [encabezados]
    for op in operaciones:
        lineas_inf = [l for l in op.lineas if l.genera_comision]
        if not lineas_inf:
            data.append([
                formato_fecha(op.fecha), op.tipo.capitalize()[:11], _corta(op.propiedad),
                op.moneda, "—", "Sin Infinity", "—", "—", "—",
            ])
            continue
        for l in lineas_inf:
            data.append([
                formato_fecha(op.fecha),
                op.tipo.capitalize()[:11],
                _corta(op.propiedad),
                op.moneda,
                l.etiqueta_rol.replace("Punta ", ""),
                _corta(l.agente.nombre_completo if l.agente else "Sin agente", 22),
                formato_moneda(l.comision_bruta, op.moneda),
                formato_moneda(l.monto_inmobiliaria, op.moneda),
                formato_moneda(l.monto_agente, op.moneda),
            ])
            for fig in l.figuras:
                data.append([
                    "", "", "", "",
                    "↳ " + fig.etiqueta_tipo,
                    _corta(fig.agente.nombre_completo if fig.agente else "Sin agente", 22),
                    formato_moneda(fig.comision_bruta, op.moneda),
                    formato_moneda(fig.monto_inmobiliaria, op.moneda),
                    formato_moneda(fig.monto_agente, op.moneda),
                ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (6, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#" + GRIS_CLARO)]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elementos.append(tabla)

    # Balance por moneda
    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph("Balance del período (por moneda)", estilo_seccion))
    balance = balance_por_moneda(operaciones)
    for moneda, datos in balance.items():
        resumen = [
            ["Moneda", "Operaciones", "Comisión bruta", "Total inmobiliaria", "Total agentes"],
            [
                moneda,
                str(datos["cantidad_operaciones"]),
                formato_moneda(datos["comision_bruta"], moneda),
                formato_moneda(datos["monto_inmobiliaria"], moneda),
                formato_moneda(datos["monto_agente_total"], moneda),
            ],
        ]
        t = Table(resumen, colWidths=[40 * mm, 40 * mm, 60 * mm, 60 * mm, 60 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NARANJA)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        elementos.append(t)
        elementos.append(Spacer(1, 4))

        if datos["por_agente"]:
            ag = [["Agente", "Monto que se llevó"]]
            for nombre, monto in datos["por_agente"]:
                ag.append([nombre, formato_moneda(monto, moneda)])
            ta = Table(ag, colWidths=[120 * mm, 60 * mm])
            ta.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + AZUL)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ]))
            elementos.append(ta)
            elementos.append(Spacer(1, 10))

    doc.build(elementos)
    salida.seek(0)
    return salida


def _corta(texto, n=26):
    texto = texto or ""
    return texto if len(texto) <= n else texto[: n - 1] + "…"


# --------------------------------------------------------------------------
# PDF — Desempeño (ranking de ingresos en pesos)
# --------------------------------------------------------------------------
def exportar_desempeno_pdf(filas, resumen, etiqueta_periodo, total_ops):
    """PDF del ranking de ingresos en pesos (agentes + inmobiliaria)."""
    salida = BytesIO()
    doc = SimpleDocTemplate(
        salida, pagesize=A4,
        leftMargin=16 * mm, rightMargin=16 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title="Desempeño — ingresos en pesos",
    )
    estilos = getSampleStyleSheet()
    est_sub = ParagraphStyle("Sub", parent=estilos["Normal"], fontSize=10,
                             textColor=colors.HexColor("#333333"), spaceAfter=3)
    est_seccion = ParagraphStyle("Sec", parent=estilos["Heading2"], fontSize=13,
                                 textColor=colors.HexColor("#" + AZUL), spaceBefore=8, spaceAfter=8)

    elementos = []
    if os.path.exists(LOGO_PDF):
        try:
            ir = ImageReader(LOGO_PDF)
            iw, ih = ir.getSize()
            alto = 18 * mm
            logo = Image(LOGO_PDF, width=alto * iw / ih, height=alto)
            logo.hAlign = "LEFT"
            elementos.append(logo)
            elementos.append(Spacer(1, 4))
        except Exception:
            pass

    elementos.append(Paragraph("Desempeño — ingresos en pesos", est_seccion))
    tc = resumen["tipo_cambio"]
    elementos.append(Paragraph(
        f"{etiqueta_periodo} · {total_ops} operación(es) · "
        f"Tipo de cambio: US$ 1 = {formato_moneda(tc, 'ARS', 0)}", est_sub))
    elementos.append(Spacer(1, 8))

    # Resumen
    resumen_data = [
        ["Comisión bruta total", "Ingresos inmobiliaria", "Ingresos agentes"],
        [
            formato_moneda(resumen["bruta"], "ARS", 0),
            formato_moneda(resumen["inmobiliaria"], "ARS", 0),
            formato_moneda(resumen["agentes"], "ARS", 0),
        ],
    ]
    tr = Table(resumen_data, colWidths=[59 * mm, 59 * mm, 59 * mm])
    tr.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
    ]))
    elementos.append(tr)
    elementos.append(Spacer(1, 12))

    # Ranking
    encabezados = ["#", "Nombre", "Puntas", "De US$ (en $)", "De $", "Total en $"]
    data = [encabezados]
    fila_inmo = None
    for i, f in enumerate(filas, 1):
        if f["tipo"] == "inmobiliaria":
            fila_inmo = i  # fila i de data (i porque encabezado es 0)
        data.append([
            str(i),
            f["nombre"] + ("  (la inmobiliaria)" if f["tipo"] == "inmobiliaria" else ""),
            str(f["puntas"]),
            formato_moneda(f["de_usd"], "ARS", 0) if f["de_usd"] else "—",
            formato_moneda(f["de_ars"], "ARS", 0) if f["de_ars"] else "—",
            formato_moneda(f["ars"], "ARS", 0),
        ])

    tabla = Table(data, colWidths=[10 * mm, 52 * mm, 16 * mm, 34 * mm, 30 * mm, 35 * mm],
                  repeatRows=1)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + AZUL)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "CENTER"),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#" + GRIS_CLARO)]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if fila_inmo is not None:
        estilo.append(("BACKGROUND", (0, fila_inmo), (-1, fila_inmo), colors.HexColor("#FCEFE3")))
        estilo.append(("FONTNAME", (0, fila_inmo), (-1, fila_inmo), "Helvetica-Bold"))
    tabla.setStyle(TableStyle(estilo))
    elementos.append(tabla)

    doc.build(elementos)
    salida.seek(0)
    return salida
